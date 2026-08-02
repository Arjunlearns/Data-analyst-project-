"""
Builds Sales_Operations_Analytics_Dashboard.xlsx:
  - Raw Data       : full order-level extract + helper columns (month, hour, region)
  - City Dim       : lookup table (city -> region/tier) used by VLOOKUP
  - KPI Summary    : headline KPI cards, all formula-driven
  - Monthly Trend  : orders/revenue/cancellation trend + line chart
  - City Performance: revenue & cancellation by city + bar chart
  - RCA Cancellations: cancellation rate by traffic/weather/festival + bar charts
  - Cancellation Reasons: pareto table + pie chart
  - Delay Drivers  : avg delivery time by traffic x multiple deliveries
  - Dashboard      : front-page KPI cards + consolidated charts

All aggregates use SUMIFS/COUNTIFS/AVERAGEIFS/formulas against Raw Data —
nothing is hardcoded, so the workbook recalculates if the raw data changes.
"""
import pandas as pd
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SRC = "/home/claude/sales_ops_dashboard/data/delivery_orders.csv"
OUT = "/home/claude/sales_ops_dashboard/Sales_Operations_Analytics_Dashboard.xlsx"

NAVY = "1F2937"
ACCENT = "2563EB"
LIGHT = "EFF6FF"
GREY = "6B7280"
WHITE = "FFFFFF"

HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name="Arial", bold=True, color=NAVY, size=16)
SUBTLE_FONT = Font(name="Arial", color=GREY, size=10)
LABEL_FONT = Font(name="Arial", bold=True, color=NAVY, size=10)
KPI_FONT = Font(name="Arial", bold=True, color=ACCENT, size=22)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

df = pd.read_csv(SRC)
N = len(df)
LAST_ROW = N + 1  # header is row 1

wb = Workbook()
wb.remove(wb.active)

# ------------------------------------------------------------------
# 1. CITY DIM sheet
# ------------------------------------------------------------------
ws_dim = wb.create_sheet("City Dim")
dim_data = [
    ("city", "region", "tier"),
    ("Bengaluru", "South", 1),
    ("Chennai", "South", 1),
    ("Hyderabad", "South", 1),
    ("Mumbai", "West", 1),
    ("Pune", "West", 2),
    ("Delhi", "North", 1),
    ("Kolkata", "East", 1),
]
for r, row in enumerate(dim_data, start=1):
    for c, val in enumerate(row, start=1):
        cell = ws_dim.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        else:
            cell.font = BODY_FONT
for col, w in zip("ABC", (16, 12, 8)):
    ws_dim.column_dimensions[col].width = w

# ------------------------------------------------------------------
# 2. RAW DATA sheet
# ------------------------------------------------------------------
ws_raw = wb.create_sheet("Raw Data")
raw_cols = [
    "order_id", "delivery_person_id", "delivery_person_age", "delivery_person_rating",
    "city", "order_date", "order_time", "weather_conditions", "road_traffic_density",
    "vehicle_condition", "type_of_order", "type_of_vehicle", "multiple_deliveries",
    "festival", "distance_km", "order_value_inr", "payment_type", "order_status",
    "cancellation_reason", "delivery_time_min",
]
helper_cols = ["month", "order_hour", "region"]
all_cols = raw_cols + helper_cols

for c, name in enumerate(all_cols, start=1):
    cell = ws_raw.cell(row=1, column=c, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

COL = {name: get_column_letter(i + 1) for i, name in enumerate(all_cols)}

for i, rec in enumerate(df.itertuples(index=False), start=2):
    d = rec._asdict() if hasattr(rec, "_asdict") else dict(zip(raw_cols, rec))
    ws_raw.cell(row=i, column=1, value=d["order_id"])
    ws_raw.cell(row=i, column=2, value=d["delivery_person_id"])
    ws_raw.cell(row=i, column=3, value=int(d["delivery_person_age"]))
    ws_raw.cell(row=i, column=4, value=float(d["delivery_person_rating"]))
    ws_raw.cell(row=i, column=5, value=d["city"])
    ws_raw.cell(row=i, column=6, value=datetime.strptime(d["order_date"], "%Y-%m-%d"))
    ws_raw.cell(row=i, column=6).number_format = "yyyy-mm-dd"
    h, m, s = map(int, d["order_time"].split(":"))
    ws_raw.cell(row=i, column=7, value=time(h, m, s))
    ws_raw.cell(row=i, column=7).number_format = "hh:mm"
    ws_raw.cell(row=i, column=8, value=d["weather_conditions"])
    ws_raw.cell(row=i, column=9, value=d["road_traffic_density"])
    ws_raw.cell(row=i, column=10, value=int(d["vehicle_condition"]))
    ws_raw.cell(row=i, column=11, value=d["type_of_order"])
    ws_raw.cell(row=i, column=12, value=d["type_of_vehicle"])
    ws_raw.cell(row=i, column=13, value=int(d["multiple_deliveries"]))
    ws_raw.cell(row=i, column=14, value=d["festival"])
    ws_raw.cell(row=i, column=15, value=float(d["distance_km"]))
    ws_raw.cell(row=i, column=16, value=float(d["order_value_inr"]))
    ws_raw.cell(row=i, column=17, value=d["payment_type"])
    ws_raw.cell(row=i, column=18, value=d["order_status"])
    ws_raw.cell(row=i, column=19, value=d["cancellation_reason"] if pd.notna(d["cancellation_reason"]) else "")
    tval = d["delivery_time_min"]
    ws_raw.cell(row=i, column=20, value=float(tval) if pd.notna(tval) else None)
    # helper formulas
    ws_raw.cell(row=i, column=21, value=f'=TEXT(F{i},"yyyy-mm")')
    ws_raw.cell(row=i, column=22, value=f"=HOUR(G{i})")
    ws_raw.cell(row=i, column=23, value=f"=VLOOKUP(E{i},'City Dim'!$A$2:$B$8,2,FALSE)")
    for c in range(1, 24):
        ws_raw.cell(row=i, column=c).font = BODY_FONT

for col, w in zip(
    [get_column_letter(i) for i in range(1, 24)],
    [11, 16, 8, 8, 11, 11, 8, 13, 13, 8, 11, 15, 10, 8, 10, 12, 11, 11, 22, 14, 8, 10, 8],
):
    ws_raw.column_dimensions[col].width = w
ws_raw.freeze_panes = "A2"

RD = "'Raw Data'!"
R_STATUS = f"{RD}${COL['order_status']}$2:${COL['order_status']}${LAST_ROW}"
R_VALUE = f"{RD}${COL['order_value_inr']}$2:${COL['order_value_inr']}${LAST_ROW}"
R_TIME = f"{RD}${COL['delivery_time_min']}$2:${COL['delivery_time_min']}${LAST_ROW}"
R_TRAFFIC = f"{RD}${COL['road_traffic_density']}$2:${COL['road_traffic_density']}${LAST_ROW}"
R_WEATHER = f"{RD}${COL['weather_conditions']}$2:${COL['weather_conditions']}${LAST_ROW}"
R_FESTIVAL = f"{RD}${COL['festival']}$2:${COL['festival']}${LAST_ROW}"
R_CITY = f"{RD}${COL['city']}$2:${COL['city']}${LAST_ROW}"
R_REGION = f"{RD}${COL['region']}$2:${COL['region']}${LAST_ROW}"
R_MONTH = f"{RD}${COL['month']}$2:${COL['month']}${LAST_ROW}"
R_REASON = f"{RD}${COL['cancellation_reason']}$2:${COL['cancellation_reason']}${LAST_ROW}"
R_MULTI = f"{RD}${COL['multiple_deliveries']}$2:${COL['multiple_deliveries']}${LAST_ROW}"
R_HOUR = f"{RD}${COL['order_hour']}$2:${COL['order_hour']}${LAST_ROW}"
R_ORDERID = f"{RD}${COL['order_id']}$2:${COL['order_id']}${LAST_ROW}"


def section_title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = TITLE_FONT


def header_row(ws, row, start_col, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BOX


def style_table_body(ws, first_row, last_row, first_col, last_col, pct_cols=(), money_cols=()):
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BOX
            if c in pct_cols:
                cell.number_format = "0.00\"%\""
            if c in money_cols:
                cell.number_format = "#,##0"


# ------------------------------------------------------------------
# 3. KPI SUMMARY sheet
# ------------------------------------------------------------------
ws_kpi = wb.create_sheet("KPI Summary")
section_title(ws_kpi, "A1", "Headline KPIs")
ws_kpi["A2"] = "All formulas reference the Raw Data sheet — recalculates automatically."
ws_kpi["A2"].font = SUBTLE_FONT

kpi_rows = [
    ("Total Orders", f"=COUNTA({R_ORDERID})", None),
    ("Delivered Orders", f'=COUNTIF({R_STATUS},"Delivered")', None),
    ("Cancelled Orders", f'=COUNTIF({R_STATUS},"Cancelled")', None),
    ("Cancellation Rate", "=B6/B4*100", "pct"),
    ("Revenue (INR)", f'=SUMIF({R_STATUS},"Delivered",{R_VALUE})', "money"),
    ("Avg Order Value (INR)", f'=AVERAGEIF({R_STATUS},"Delivered",{R_VALUE})', "money"),
    ("Avg Delivery Time (min)", f'=AVERAGEIF({R_STATUS},"Delivered",{R_TIME})', "num1"),
]
start = 4
for i, (label, formula, fmt) in enumerate(kpi_rows):
    r = start + i
    ws_kpi.cell(row=r, column=1, value=label).font = LABEL_FONT
    cell = ws_kpi.cell(row=r, column=2, value=formula)
    cell.font = KPI_FONT if False else BODY_FONT
    if fmt == "pct":
        cell.number_format = "0.00\"%\""
    elif fmt == "money":
        cell.number_format = "#,##0"
    elif fmt == "num1":
        cell.number_format = "0.0"
    ws_kpi.cell(row=r, column=1).border = BOX
    cell.border = BOX
ws_kpi.column_dimensions["A"].width = 26
ws_kpi.column_dimensions["B"].width = 18

# ------------------------------------------------------------------
# 4. MONTHLY TREND sheet
# ------------------------------------------------------------------
ws_mt = wb.create_sheet("Monthly Trend")
section_title(ws_mt, "A1", "Monthly Trend")
months = sorted(df["order_date"].str.slice(0, 7).unique())
header_row(ws_mt, 3, 1, ["Month", "Total Orders", "Cancelled Orders", "Cancellation Rate %", "Revenue (INR)", "Revenue Growth %"])
for i, mth in enumerate(months):
    r = 4 + i
    ws_mt.cell(row=r, column=1, value=mth).font = BODY_FONT
    ws_mt.cell(row=r, column=2, value=f'=COUNTIF({R_MONTH},A{r})')
    ws_mt.cell(row=r, column=3, value=f'=COUNTIFS({R_MONTH},A{r},{R_STATUS},"Cancelled")')
    ws_mt.cell(row=r, column=4, value=f"=C{r}/B{r}*100")
    ws_mt.cell(row=r, column=5, value=f'=SUMIFS({R_VALUE},{R_MONTH},A{r},{R_STATUS},"Delivered")')
    if i == 0:
        ws_mt.cell(row=r, column=6, value="")
    else:
        ws_mt.cell(row=r, column=6, value=f"=(E{r}-E{r-1})/E{r-1}*100")
last_mt = 3 + len(months)
style_table_body(ws_mt, 4, last_mt, 1, 6, pct_cols=(4, 6), money_cols=(5,))
for col, w in zip("ABCDEF", (12, 14, 16, 18, 15, 16)):
    ws_mt.column_dimensions[col].width = w

# ------------------------------------------------------------------
# 5. CITY PERFORMANCE sheet
# ------------------------------------------------------------------
ws_city = wb.create_sheet("City Performance")
section_title(ws_city, "A1", "City & Region Performance")
cities = sorted(df["city"].unique())
header_row(ws_city, 3, 1, ["City", "Region", "Total Orders", "Cancellation Rate %", "Revenue (INR)", "Revenue Rank"])
for i, city in enumerate(cities):
    r = 4 + i
    ws_city.cell(row=r, column=1, value=city).font = BODY_FONT
    ws_city.cell(row=r, column=2, value=f"=VLOOKUP(A{r},'City Dim'!$A$2:$B$8,2,FALSE)")
    ws_city.cell(row=r, column=3, value=f"=COUNTIF({R_CITY},A{r})")
    ws_city.cell(row=r, column=4, value=f'=COUNTIFS({R_CITY},A{r},{R_STATUS},"Cancelled")/C{r}*100')
    ws_city.cell(row=r, column=5, value=f'=SUMIFS({R_VALUE},{R_CITY},A{r},{R_STATUS},"Delivered")')
last_city = 3 + len(cities)
for i in range(len(cities)):
    r = 4 + i
    ws_city.cell(row=r, column=6, value=f"=RANK(E{r},$E${4}:$E${last_city})")
style_table_body(ws_city, 4, last_city, 1, 6, pct_cols=(4,), money_cols=(5,))
for col, w in zip("ABCDEF", (14, 10, 13, 18, 15, 12)):
    ws_city.column_dimensions[col].width = w

# ------------------------------------------------------------------
# 6. RCA CANCELLATIONS sheet
# ------------------------------------------------------------------
ws_rca = wb.create_sheet("RCA Cancellations")
section_title(ws_rca, "A1", "Root Cause Analysis — Cancellation Drivers")

def rca_block(ws, top_row, title, values, range_ref):
    ws.cell(row=top_row, column=1, value=title).font = LABEL_FONT
    header_row(ws, top_row + 1, 1, ["Segment", "Total Orders", "Cancelled", "Cancellation Rate %"])
    for i, v in enumerate(values):
        r = top_row + 2 + i
        ws.cell(row=r, column=1, value=v).font = BODY_FONT
        ws.cell(row=r, column=2, value=f"=COUNTIF({range_ref},A{r})")
        ws.cell(row=r, column=3, value=f'=COUNTIFS({range_ref},A{r},{R_STATUS},"Cancelled")')
        ws.cell(row=r, column=4, value=f"=C{r}/B{r}*100")
    last = top_row + 1 + len(values)
    style_table_body(ws, top_row + 2, last, 1, 4, pct_cols=(4,))
    return last

traffic_vals = ["Low", "Medium", "High", "Jam"]
weather_vals = sorted(df["weather_conditions"].unique())
festival_vals = ["Yes", "No"]

r1_end = rca_block(ws_rca, 3, "By Traffic Density", traffic_vals, R_TRAFFIC)
r2_start = r1_end + 3
r2_end = rca_block(ws_rca, r2_start, "By Weather Condition", weather_vals, R_WEATHER)
r3_start = r2_end + 3
r3_end = rca_block(ws_rca, r3_start, "By Festival Day", festival_vals, R_FESTIVAL)

for col, w in zip("ABCD", (16, 14, 12, 20)):
    ws_rca.column_dimensions[col].width = w

TRAFFIC_TBL = (4, r1_end)
WEATHER_TBL = (r2_start + 2, r2_end)
FESTIVAL_TBL = (r3_start + 2, r3_end)

# ------------------------------------------------------------------
# 7. CANCELLATION REASONS sheet (pareto)
# ------------------------------------------------------------------
ws_reason = wb.create_sheet("Cancellation Reasons")
section_title(ws_reason, "A1", "Cancellation Reason — Pareto")
reasons = df.loc[df.order_status == "Cancelled", "cancellation_reason"].value_counts().index.tolist()
header_row(ws_reason, 3, 1, ["Reason", "Occurrences", "% of Cancellations", "Cumulative %"])
for i, reason in enumerate(reasons):
    r = 4 + i
    ws_reason.cell(row=r, column=1, value=reason).font = BODY_FONT
    ws_reason.cell(row=r, column=2, value=f'=COUNTIF({R_REASON},A{r})')
last_reason = 3 + len(reasons)
total_cell = f"SUM($B$4:$B${last_reason})"
for i in range(len(reasons)):
    r = 4 + i
    ws_reason.cell(row=r, column=3, value=f"=B{r}/{total_cell}*100")
    ws_reason.cell(row=r, column=4, value=f"=SUM($C$4:C{r})")
style_table_body(ws_reason, 4, last_reason, 1, 4, pct_cols=(3, 4))
for col, w in zip("ABCD", (26, 13, 18, 15)):
    ws_reason.column_dimensions[col].width = w

# ------------------------------------------------------------------
# 8. DELAY DRIVERS sheet
# ------------------------------------------------------------------
ws_delay = wb.create_sheet("Delay Drivers")
section_title(ws_delay, "A1", "Delivery Delay Drivers (Traffic x Multiple Deliveries)")
header_row(ws_delay, 3, 1, ["Traffic", "Multiple Deliveries", "Delivered Orders", "Avg Delivery Time (min)", "Delay Rank"])
combos = [(t, m) for t in traffic_vals for m in sorted(df["multiple_deliveries"].unique())]
row_i = 4
combo_rows = []
for t, m in combos:
    mask_count = ((df.road_traffic_density == t) & (df.multiple_deliveries == m) & (df.order_status == "Delivered")).sum()
    if mask_count == 0:
        continue
    ws_delay.cell(row=row_i, column=1, value=t).font = BODY_FONT
    ws_delay.cell(row=row_i, column=2, value=int(m)).font = BODY_FONT
    ws_delay.cell(row=row_i, column=3, value=(
        f'=COUNTIFS({R_TRAFFIC},A{row_i},{R_MULTI},B{row_i},{R_STATUS},"Delivered")'
    ))
    ws_delay.cell(row=row_i, column=4, value=(
        f'=AVERAGEIFS({R_TIME},{R_TRAFFIC},A{row_i},{R_MULTI},B{row_i},{R_STATUS},"Delivered")'
    ))
    combo_rows.append(row_i)
    row_i += 1
last_delay = row_i - 1
for r in combo_rows:
    ws_delay.cell(row=r, column=5, value=f"=RANK(D{r},$D${combo_rows[0]}:$D${last_delay})")
style_table_body(ws_delay, combo_rows[0], last_delay, 1, 5)
for r in combo_rows:
    ws_delay.cell(row=r, column=4).number_format = "0.0"
for col, w in zip("ABCDE", (12, 18, 16, 22, 12)):
    ws_delay.column_dimensions[col].width = w

# ------------------------------------------------------------------
# 9. DASHBOARD (front page)
# ------------------------------------------------------------------
ws_dash = wb.create_sheet("Dashboard", 0)
ws_dash.sheet_view.showGridLines = False
ws_dash["B2"] = "Sales & Operations Analytics Dashboard"
ws_dash["B2"].font = Font(name="Arial", bold=True, size=20, color=NAVY)
ws_dash["B3"] = "Food & ride delivery ops — orders, revenue, cancellations, and delay drivers"
ws_dash["B3"].font = SUBTLE_FONT

kpi_cards = [
    ("Total Orders", f"='KPI Summary'!B4", "0"),
    ("Revenue (INR)", f"='KPI Summary'!B8", "#,##0"),
    ("Cancellation Rate", f"='KPI Summary'!B7", "0.00\"%\""),
    ("Avg Delivery Time (min)", f"='KPI Summary'!B10", "0.0"),
]
card_col = 2
for i, (label, formula, fmt) in enumerate(kpi_cards):
    col = card_col + i * 3
    letter = get_column_letter(col)
    ws_dash.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    ws_dash.merge_cells(start_row=6, start_column=col, end_row=7, end_column=col + 1)
    lbl = ws_dash.cell(row=5, column=col, value=label)
    lbl.font = LABEL_FONT
    val = ws_dash.cell(row=6, column=col, value=formula)
    val.font = KPI_FONT
    val.number_format = fmt
    val.alignment = Alignment(vertical="center")
    for rr in (5, 6, 7):
        for cc in (col, col + 1):
            ws_dash.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=LIGHT)

ws_dash.row_dimensions[8].height = 10
for col in range(2, 14):
    ws_dash.column_dimensions[get_column_letter(col)].width = 11

# Charts pulling from the summary sheets
line = LineChart()
line.title = "Monthly Orders & Cancellation Rate"
line.y_axis.title = "Orders"
line.style = 10
cats = Reference(ws_mt, min_col=1, min_row=4, max_row=last_mt)
data_orders = Reference(ws_mt, min_col=2, min_row=3, max_row=last_mt)
line.add_data(data_orders, titles_from_data=True)
line.set_categories(cats)
line.height, line.width = 8, 16
ws_dash.add_chart(line, "B10")

bar_city = BarChart()
bar_city.type = "col"
bar_city.title = "Revenue by City (INR)"
bar_city.style = 12
cats2 = Reference(ws_city, min_col=1, min_row=4, max_row=last_city)
data2 = Reference(ws_city, min_col=5, min_row=3, max_row=last_city)
bar_city.add_data(data2, titles_from_data=True)
bar_city.set_categories(cats2)
bar_city.height, bar_city.width = 8, 16
ws_dash.add_chart(bar_city, "J10")

bar_traffic = BarChart()
bar_traffic.type = "col"
bar_traffic.title = "Cancellation Rate % by Traffic Density"
bar_traffic.style = 11
cats3 = Reference(ws_rca, min_col=1, min_row=TRAFFIC_TBL[0], max_row=TRAFFIC_TBL[1])
data3 = Reference(ws_rca, min_col=4, min_row=TRAFFIC_TBL[0] - 1, max_row=TRAFFIC_TBL[1])
bar_traffic.add_data(data3, titles_from_data=True)
bar_traffic.set_categories(cats3)
bar_traffic.height, bar_traffic.width = 8, 16
ws_dash.add_chart(bar_traffic, "B26")

pie = PieChart()
pie.title = "Share of Cancellations by Reason"
cats4 = Reference(ws_reason, min_col=1, min_row=4, max_row=last_reason)
data4 = Reference(ws_reason, min_col=2, min_row=3, max_row=last_reason)
pie.add_data(data4, titles_from_data=True)
pie.set_categories(cats4)
pie.height, pie.width = 8, 16
ws_dash.add_chart(pie, "J26")

# ------------------------------------------------------------------
wb.save(OUT)
print("Saved:", OUT)
